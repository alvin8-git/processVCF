#!/usr/bin/env python3
"""
Excel to HTML Variant Report Converter

Converts variant Excel files from processVCF.sh output to interactive HTML reports.
Creates a dashboard landing page and individual variant pages.

Usage:
    python3 excel_to_html_report.py <input_excel_file> [output_directory]

Author: Clinical Genomics Pipeline
"""

import os
import sys
import re
import json
from pathlib import Path
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from html import escape

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


@dataclass
class ColumnGroups:
    """Column groupings for variant data"""
    BASIC_INFO = list(range(1, 19))          # Col 1-18: Basic variant information
    SAMPLE_COMPARISON = list(range(19, 23))   # Col 19-22: Sample comparison data
    ADDITIONAL_INFO = list(range(23, 32))     # Col 23-31: Additional variant information
    POPULATION_DB = list(range(32, 90))       # Col 32-89: Population databases
    COMPUTATIONAL = list(range(90, 155))      # Col 90-154: Computational scores


# Population database groupings
POP_DB_GROUPS = {
    "Singapore/Asian (SG10K)": [32, 33, 34, 35, 36, 37, 38],  # AF_All to SAS_AF
    "ESP": [39],  # esp6500siv2_all
    "ExAC": list(range(40, 48)),  # ExAC_ALL to ExAC_SAS
    "1000 Genomes": list(range(48, 54)),  # 1000g2015aug_*
    "Kaviar": [54, 55, 56],  # Kaviar_AF, AC, AN
    "gnomAD Exome": list(range(57, 66)),  # AF to AF_oth
    "gnomAD Genome": list(range(66, 74)),  # gnomAD_genome_*
    "HRC": list(range(74, 80)),  # HRC_*
    "GME": list(range(80, 88)),  # GME_*
    "Other": [88, 89],  # cg69, nci60
}

# ACMG criteria columns
ACMG_CRITERIA = {
    "Very Strong Pathogenic": [(91, "PVS1")],
    "Strong Pathogenic": [(92, "PS1"), (93, "PS2"), (94, "PS3"), (95, "PS4")],
    "Moderate Pathogenic": [(96, "PM1"), (97, "PM2"), (98, "PM3"), (99, "PM4"), (100, "PM5"), (101, "PM6")],
    "Supporting Pathogenic": [(102, "PP1"), (103, "PP2"), (104, "PP3"), (105, "PP4"), (106, "PP5")],
    "Stand-Alone Benign": [(107, "BA1")],
    "Strong Benign": [(108, "BS1"), (109, "BS2"), (110, "BS3"), (111, "BS4")],
    "Supporting Benign": [(112, "BP1"), (113, "BP2"), (114, "BP3"), (115, "BP4"), (116, "BP5"), (117, "BP6"), (118, "BP7")],
}

# ClinVar columns
CLINVAR_COLS = [119, 120, 121, 122, 123]

# Prediction score columns
PREDICTION_SCORES = {
    "InterVar": [(90, "InterVar_automated")],
    "M-CAP": [(124, "MCAP")],
    "REVEL": [(125, "REVEL")],
    "SIFT": [(126, "SIFT_score"), (127, "SIFT_pred")],
    "PolyPhen2 HDIV": [(128, "Polyphen2_HDIV_score"), (129, "Polyphen2_HDIV_pred")],
    "PolyPhen2 HVAR": [(130, "Polyphen2_HVAR_score"), (131, "Polyphen2_HVAR_pred")],
    "LRT": [(132, "LRT_score"), (133, "LRT_pred")],
    "MutationTaster": [(134, "MutationTaster_score"), (135, "MutationTaster_pred")],
    "MutationAssessor": [(136, "MutationAssessor_score"), (137, "MutationAssessor_pred")],
    "FATHMM": [(138, "FATHMM_score"), (139, "FATHMM_pred")],
    "RadialSVM": [(140, "RadialSVM_score"), (141, "RadialSVM_pred")],
    "LR": [(142, "LR_score"), (143, "LR_pred")],
    "VEST3": [(144, "VEST3_score")],
    "CADD": [(145, "CADD_raw"), (146, "CADD_phred")],
    "GERP++": [(147, "GERP++_RS")],
    "phyloP": [(148, "phyloP46way_placental"), (149, "phyloP100way_vertebrate")],
    "SiPhy": [(150, "SiPhy_29way_logOdds")],
    "Conservation": [(151, "phastConsElements46way"), (152, "phastConsElements100way"), (153, "tfbsConsSites"), (154, "targetScanS")],
}


def get_css_styles() -> str:
    """Return CSS styles for the HTML reports"""
    return """
:root {
    --primary-color: #2c3e50;
    --secondary-color: #3498db;
    --accent-color: #e74c3c;
    --success-color: #27ae60;
    --warning-color: #f39c12;
    --light-bg: #f8f9fa;
    --border-color: #dee2e6;
    --text-muted: #6c757d;
    --tier1-color: #dc3545;
    --tier2-color: #fd7e14;
    --tier3-color: #ffc107;
    --tier4-color: #28a745;
    --pathogenic-color: #dc3545;
    --likely-pathogenic-color: #fd7e14;
    --vus-color: #6c757d;
    --likely-benign-color: #17a2b8;
    --benign-color: #28a745;
}

* {
    box-sizing: border-box;
}

body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
    line-height: 1.5;
    color: #212529;
    background-color: #f5f6fa;
    margin: 0;
    padding: 0;
}

.container {
    max-width: 1600px;
    margin: 0 auto;
    padding: 20px;
}

/* Header Styles */
.header {
    background: linear-gradient(135deg, var(--primary-color) 0%, #34495e 100%);
    color: white;
    padding: 20px 30px;
    margin-bottom: 20px;
    border-radius: 8px;
    box-shadow: 0 4px 6px rgba(0,0,0,0.1);
}

.header h1 {
    margin: 0 0 10px 0;
    font-size: 1.8rem;
    font-weight: 600;
}

.header .subtitle {
    opacity: 0.9;
    font-size: 0.95rem;
}

/* Navigation */
.breadcrumb {
    background: white;
    padding: 12px 20px;
    border-radius: 6px;
    margin-bottom: 20px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
}

.breadcrumb a {
    color: var(--secondary-color);
    text-decoration: none;
}

.breadcrumb a:hover {
    text-decoration: underline;
}

/* Gene Hero Section */
.gene-hero {
    background: white;
    border-radius: 10px;
    padding: 25px 30px;
    margin-bottom: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    display: flex;
    flex-wrap: wrap;
    align-items: center;
    gap: 20px;
}

.gene-name {
    font-size: 2.8rem;
    font-weight: 700;
    color: var(--primary-color);
    margin: 0;
}

.variant-notation {
    font-size: 1.4rem;
    color: var(--text-muted);
    font-family: 'Monaco', 'Menlo', monospace;
}

.classification-badge {
    display: inline-block;
    padding: 8px 20px;
    border-radius: 25px;
    font-weight: 600;
    font-size: 0.95rem;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

.tier-i { background-color: var(--tier1-color); color: white; }
.tier-ii { background-color: var(--tier2-color); color: white; }
.tier-iii { background-color: var(--tier3-color); color: #212529; }
.tier-iv { background-color: var(--tier4-color); color: white; }

.pathogenic { background-color: var(--pathogenic-color); color: white; }
.likely-pathogenic { background-color: var(--likely-pathogenic-color); color: white; }
.vus { background-color: var(--vus-color); color: white; }
.likely-benign { background-color: var(--likely-benign-color); color: white; }
.benign { background-color: var(--benign-color); color: white; }

/* Panel Styles */
.panel {
    background: white;
    border-radius: 10px;
    margin-bottom: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    overflow: hidden;
}

.panel-header {
    background: var(--light-bg);
    padding: 15px 20px;
    border-bottom: 1px solid var(--border-color);
    cursor: pointer;
    display: flex;
    justify-content: space-between;
    align-items: center;
    transition: background-color 0.2s;
}

.panel-header:hover {
    background: #e9ecef;
}

.panel-title {
    font-size: 1.1rem;
    font-weight: 600;
    color: var(--primary-color);
    margin: 0;
}

.panel-toggle {
    font-size: 1.2rem;
    color: var(--text-muted);
    transition: transform 0.3s;
}

.panel.collapsed .panel-toggle {
    transform: rotate(-90deg);
}

.panel.collapsed .panel-content {
    display: none;
}

.panel-content {
    padding: 20px;
}

/* Data Grid */
.data-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
    gap: 15px;
}

.data-item {
    padding: 10px 15px;
    background: var(--light-bg);
    border-radius: 6px;
    border-left: 3px solid var(--secondary-color);
}

.data-label {
    font-size: 0.75rem;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 4px;
}

.data-value {
    font-size: 0.95rem;
    color: #212529;
    word-break: break-word;
}

.data-value.large {
    font-size: 1.2rem;
    font-weight: 600;
}

.data-value.monospace {
    font-family: 'Monaco', 'Menlo', monospace;
    font-size: 0.9rem;
}

/* Population Frequency Bars */
.freq-bar-container {
    width: 100%;
    height: 6px;
    background: #e9ecef;
    border-radius: 3px;
    margin-top: 5px;
    overflow: hidden;
}

.freq-bar {
    height: 100%;
    background: var(--secondary-color);
    border-radius: 3px;
    transition: width 0.3s;
}

.freq-bar.rare { background: var(--success-color); }
.freq-bar.low { background: var(--warning-color); }
.freq-bar.common { background: var(--accent-color); }

/* ACMG Criteria Chips */
.acmg-grid {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-bottom: 15px;
}

.acmg-chip {
    display: inline-flex;
    align-items: center;
    padding: 6px 12px;
    border-radius: 20px;
    font-size: 0.8rem;
    font-weight: 600;
}

.acmg-chip.inactive {
    background: #e9ecef;
    color: #adb5bd;
}

.acmg-chip.pvs { background: #721c24; color: white; }
.acmg-chip.ps { background: var(--pathogenic-color); color: white; }
.acmg-chip.pm { background: var(--likely-pathogenic-color); color: white; }
.acmg-chip.pp { background: #f8d7da; color: #721c24; }
.acmg-chip.ba { background: #155724; color: white; }
.acmg-chip.bs { background: var(--benign-color); color: white; }
.acmg-chip.bp { background: #d4edda; color: #155724; }

/* Prediction Score Table */
.score-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.9rem;
}

.score-table th,
.score-table td {
    padding: 10px 12px;
    text-align: left;
    border-bottom: 1px solid var(--border-color);
}

.score-table th {
    background: var(--light-bg);
    font-weight: 600;
    color: var(--primary-color);
}

.score-table tr:hover {
    background: #f8f9fa;
}

.pred-badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 12px;
    font-size: 0.75rem;
    font-weight: 600;
    text-transform: uppercase;
}

.pred-d, .pred-deleterious { background: #f8d7da; color: #721c24; }
.pred-t, .pred-tolerated { background: #d4edda; color: #155724; }
.pred-p, .pred-possibly { background: #fff3cd; color: #856404; }
.pred-b, .pred-benign { background: #d4edda; color: #155724; }
.pred-n, .pred-neutral { background: #e9ecef; color: #495057; }

/* Population DB Sub-panels */
.pop-db-section {
    margin-bottom: 20px;
}

.pop-db-title {
    font-size: 0.9rem;
    font-weight: 600;
    color: var(--primary-color);
    margin-bottom: 10px;
    padding-bottom: 5px;
    border-bottom: 1px solid var(--border-color);
}

/* Dashboard Styles */
.dashboard-stats {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 15px;
    margin-bottom: 25px;
}

.stat-card {
    background: white;
    border-radius: 10px;
    padding: 20px;
    text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    transition: transform 0.2s, box-shadow 0.2s;
}

.stat-card.clickable {
    cursor: pointer;
}

.stat-card.clickable:hover {
    transform: translateY(-3px);
    box-shadow: 0 6px 20px rgba(0,0,0,0.15);
}

.stat-card.clickable.active {
    box-shadow: 0 0 0 3px var(--secondary-color);
}

.stat-value {
    font-size: 2.5rem;
    font-weight: 700;
    color: var(--primary-color);
}

.stat-label {
    font-size: 0.85rem;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

/* Panel Color Themes */
.panel.panel-basic {
    border-left: 4px solid #3498db;
}
.panel.panel-basic .panel-header {
    background: linear-gradient(135deg, #3498db 0%, #2980b9 100%);
    color: white;
}
.panel.panel-basic .panel-header:hover {
    background: linear-gradient(135deg, #2980b9 0%, #1f6dad 100%);
}
.panel.panel-basic .panel-title {
    color: white;
    font-weight: 700;
}
.panel.panel-basic .panel-toggle {
    color: rgba(255,255,255,0.8);
}
.panel.panel-basic .panel-content {
    background: linear-gradient(180deg, #ebf5fb 0%, #d6eaf8 100%);
}

.panel.panel-comparison {
    border-left: 4px solid #27ae60;
}
.panel.panel-comparison .panel-header {
    background: linear-gradient(135deg, #27ae60 0%, #1e8449 100%);
    color: white;
}
.panel.panel-comparison .panel-header:hover {
    background: linear-gradient(135deg, #1e8449 0%, #196f3d 100%);
}
.panel.panel-comparison .panel-title {
    color: white;
    font-weight: 700;
}
.panel.panel-comparison .panel-toggle {
    color: rgba(255,255,255,0.8);
}
.panel.panel-comparison .panel-content {
    background: linear-gradient(180deg, #e9f7ef 0%, #d4efdf 100%);
}

.panel.panel-additional {
    border-left: 4px solid #546e7a;
}
.panel.panel-additional .panel-header {
    background: linear-gradient(135deg, #607d8b 0%, #546e7a 100%);
    color: white;
}
.panel.panel-additional .panel-header:hover {
    background: linear-gradient(135deg, #546e7a 0%, #455a64 100%);
}
.panel.panel-additional .panel-title {
    color: white;
    font-weight: 700;
}
.panel.panel-additional .panel-toggle {
    color: rgba(255,255,255,0.8);
}
.panel.panel-additional .panel-content {
    background: linear-gradient(180deg, #eceff1 0%, #cfd8dc 100%);
}

.panel.panel-population {
    border-left: 4px solid #5c6bc0;
}
.panel.panel-population .panel-header {
    background: linear-gradient(135deg, #5c6bc0 0%, #3f51b5 100%);
    color: white;
}
.panel.panel-population .panel-header:hover {
    background: linear-gradient(135deg, #3f51b5 0%, #303f9f 100%);
}
.panel.panel-population .panel-title {
    color: white;
    font-weight: 700;
}
.panel.panel-population .panel-toggle {
    color: rgba(255,255,255,0.8);
}
.panel.panel-population .panel-content {
    background: linear-gradient(180deg, #e8eaf6 0%, #c5cae9 100%);
}

.panel.panel-computational {
    border-left: 4px solid #e67e22;
}
.panel.panel-computational .panel-header {
    background: linear-gradient(135deg, #e67e22 0%, #ca6f1e 100%);
    color: white;
}
.panel.panel-computational .panel-header:hover {
    background: linear-gradient(135deg, #ca6f1e 0%, #b9770e 100%);
}
.panel.panel-computational .panel-title {
    color: white;
    font-weight: 700;
}
.panel.panel-computational .panel-toggle {
    color: rgba(255,255,255,0.8);
}
.panel.panel-computational .panel-content {
    background: linear-gradient(180deg, #fef5e7 0%, #fdebd0 100%);
}

.panel.panel-igv {
    border-left: 4px solid #8e44ad;
}
.panel.panel-igv .panel-header {
    background: linear-gradient(135deg, #9b59b6 0%, #8e44ad 100%);
    color: white;
}
.panel.panel-igv .panel-header:hover {
    background: linear-gradient(135deg, #8e44ad 0%, #7d3c98 100%);
}
.panel.panel-igv .panel-title {
    color: white;
    font-weight: 700;
}
.panel.panel-igv .panel-toggle {
    color: rgba(255,255,255,0.8);
}
.panel.panel-igv .panel-content {
    background: linear-gradient(180deg, #f5eef8 0%, #e8daef 100%);
    text-align: center;
}
.panel.panel-igv .igv-image {
    max-width: 100%;
    height: auto;
    border-radius: 8px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    margin: 10px 0;
}
.panel.panel-igv .igv-no-image {
    color: var(--text-muted);
    font-style: italic;
    padding: 40px 20px;
}

/* Filter Bar */
.filter-bar {
    background: white;
    border-radius: 10px;
    padding: 20px;
    margin-bottom: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    display: flex;
    flex-wrap: wrap;
    gap: 15px;
    align-items: center;
}

.filter-group {
    display: flex;
    flex-direction: column;
    gap: 5px;
}

.filter-label {
    font-size: 0.75rem;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

.filter-input,
.filter-select {
    padding: 8px 12px;
    border: 1px solid var(--border-color);
    border-radius: 6px;
    font-size: 0.9rem;
    min-width: 180px;
}

.filter-input:focus,
.filter-select:focus {
    outline: none;
    border-color: var(--secondary-color);
    box-shadow: 0 0 0 3px rgba(52, 152, 219, 0.1);
}

/* Variant Table */
.variant-table {
    width: 100%;
    border-collapse: collapse;
    background: white;
    border-radius: 10px;
    overflow: hidden;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}

.variant-table th,
.variant-table td {
    padding: 14px 16px;
    text-align: left;
    border-bottom: 1px solid var(--border-color);
}

.variant-table th {
    background: var(--primary-color);
    color: white;
    font-weight: 600;
    font-size: 0.85rem;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

.variant-table tr:hover {
    background: #f8f9fa;
}

.variant-table td.gene {
    font-weight: 600;
    color: var(--primary-color);
}

.variant-table a {
    color: var(--secondary-color);
    text-decoration: none;
}

.variant-table a:hover {
    text-decoration: underline;
}

/* Responsive */
@media (max-width: 768px) {
    .container {
        padding: 10px;
    }

    .gene-hero {
        flex-direction: column;
        align-items: flex-start;
    }

    .gene-name {
        font-size: 2rem;
    }

    .data-grid {
        grid-template-columns: 1fr;
    }

    .filter-bar {
        flex-direction: column;
    }

    .filter-input,
    .filter-select {
        width: 100%;
    }
}

/* Print Styles */
@media print {
    .panel.collapsed .panel-content {
        display: block !important;
    }

    .breadcrumb,
    .filter-bar {
        display: none;
    }

    .panel {
        break-inside: avoid;
    }
}
"""


def get_javascript() -> str:
    """Return JavaScript for interactivity"""
    return """
// Panel toggle functionality
document.querySelectorAll('.panel-header').forEach(header => {
    header.addEventListener('click', () => {
        const panel = header.parentElement;
        panel.classList.toggle('collapsed');
    });
});

// Dashboard filtering (for dashboard page)
function filterVariants() {
    const searchTerm = document.getElementById('searchInput')?.value.toLowerCase() || '';
    const geneFilter = document.getElementById('geneFilter')?.value || '';
    const tierFilter = document.getElementById('tierFilter')?.value || '';

    document.querySelectorAll('.variant-row').forEach(row => {
        const gene = row.dataset.gene?.toLowerCase() || '';
        const tier = row.dataset.tier || '';
        const text = row.textContent.toLowerCase();

        const matchesSearch = text.includes(searchTerm);
        const matchesGene = !geneFilter || gene === geneFilter.toLowerCase();
        const matchesTier = !tierFilter || tier === tierFilter;

        row.style.display = (matchesSearch && matchesGene && matchesTier) ? '' : 'none';
    });

    updateStats();
}

function updateStats() {
    const visibleRows = document.querySelectorAll('.variant-row:not([style*="display: none"])').length;
    const visibleCount = document.getElementById('visibleCount');
    if (visibleCount) {
        visibleCount.textContent = visibleRows;
    }
}

// Stat card click to filter by tier
function filterByTier(tier) {
    const tierFilter = document.getElementById('tierFilter');
    if (tierFilter) {
        // Toggle: if already selected, clear filter; otherwise set it
        if (tierFilter.value === tier) {
            tierFilter.value = '';
        } else {
            tierFilter.value = tier;
        }
        filterVariants();
        updateActiveCards();
    }
}

// Clear all filters
function clearAllFilters() {
    const searchInput = document.getElementById('searchInput');
    const geneFilter = document.getElementById('geneFilter');
    const tierFilter = document.getElementById('tierFilter');

    if (searchInput) searchInput.value = '';
    if (geneFilter) geneFilter.value = '';
    if (tierFilter) tierFilter.value = '';

    filterVariants();
    updateActiveCards();
}

// Focus on gene filter dropdown
function focusGeneFilter() {
    const geneFilter = document.getElementById('geneFilter');
    if (geneFilter) {
        geneFilter.focus();
        // Open the dropdown by simulating a click (works in most browsers)
        geneFilter.click();
    }
}

// Update active state on stat cards
function updateActiveCards() {
    const tierFilter = document.getElementById('tierFilter');
    document.querySelectorAll('.stat-card.clickable').forEach(card => {
        card.classList.remove('active');
    });
    if (tierFilter && tierFilter.value) {
        const activeCard = document.querySelector(`.stat-card[data-tier="${tierFilter.value}"]`);
        if (activeCard) activeCard.classList.add('active');
    }
}

// Initialize filters on page load
document.addEventListener('DOMContentLoaded', () => {
    const searchInput = document.getElementById('searchInput');
    const geneFilter = document.getElementById('geneFilter');
    const tierFilter = document.getElementById('tierFilter');

    if (searchInput) searchInput.addEventListener('input', filterVariants);
    if (geneFilter) geneFilter.addEventListener('change', filterVariants);
    if (tierFilter) tierFilter.addEventListener('change', filterVariants);

    // Stat card click handlers
    document.querySelectorAll('.stat-card.clickable').forEach(card => {
        card.addEventListener('click', () => {
            const tier = card.dataset.tier;
            const action = card.dataset.action;

            if (tier) {
                filterByTier(tier);
            } else if (action === 'clear-filters') {
                clearAllFilters();
            } else if (action === 'focus-gene') {
                focusGeneFilter();
            }
        });
    });
});
"""


def parse_cancervar(cancervar_str: str) -> tuple:
    """Parse CancerVar classification string"""
    if not cancervar_str or cancervar_str == '.':
        return ('Unknown', 'unknown')

    cancervar_str = str(cancervar_str)

    if 'Tier_I_strong' in cancervar_str:
        return ('Tier I - Strong', 'tier-i')
    elif 'Tier_II_potential' in cancervar_str:
        return ('Tier II - Potential', 'tier-ii')
    elif 'Tier_III' in cancervar_str:
        return ('Tier III - Unknown', 'tier-iii')
    elif 'Tier_IV' in cancervar_str:
        return ('Tier IV - Benign', 'tier-iv')
    else:
        return (cancervar_str[:30], 'vus')


def parse_clinsig(clinsig_str: str) -> tuple:
    """Parse ClinVar clinical significance"""
    if not clinsig_str or clinsig_str == '.':
        return ('Not reported', 'vus')

    clinsig_lower = str(clinsig_str).lower()

    if 'pathogenic' in clinsig_lower and 'likely' not in clinsig_lower:
        return ('Pathogenic', 'pathogenic')
    elif 'likely_pathogenic' in clinsig_lower or 'likely pathogenic' in clinsig_lower:
        return ('Likely Pathogenic', 'likely-pathogenic')
    elif 'benign' in clinsig_lower and 'likely' not in clinsig_lower:
        return ('Benign', 'benign')
    elif 'likely_benign' in clinsig_lower or 'likely benign' in clinsig_lower:
        return ('Likely Benign', 'likely-benign')
    elif 'uncertain' in clinsig_lower or 'vus' in clinsig_lower:
        return ('VUS', 'vus')
    else:
        return (str(clinsig_str)[:30], 'vus')


def format_value(value: Any) -> str:
    """Format a cell value for display"""
    if value is None or value == '.' or value == '':
        return '-'
    if isinstance(value, float):
        if value < 0.0001 and value > 0:
            return f"{value:.2e}"
        elif value < 1:
            return f"{value:.6f}".rstrip('0').rstrip('.')
        else:
            return f"{value:.4f}".rstrip('0').rstrip('.')
    return str(value)


def is_empty_value(value: Any) -> bool:
    """Check if a value is empty/null"""
    return value is None or value == '.' or value == '' or str(value).strip() == ''


def get_freq_bar_class(freq: float) -> str:
    """Get CSS class for frequency bar based on frequency value"""
    if freq < 0.001:
        return 'rare'
    elif freq < 0.01:
        return 'low'
    else:
        return 'common'


def get_prediction_class(pred: str) -> str:
    """Get CSS class for prediction badge"""
    if not pred or pred == '.':
        return ''
    pred_lower = str(pred).lower()
    if pred_lower.startswith('d') or 'damaging' in pred_lower or 'deleterious' in pred_lower:
        return 'pred-d'
    elif pred_lower.startswith('t') or 'tolerated' in pred_lower:
        return 'pred-t'
    elif pred_lower.startswith('p') or 'possibly' in pred_lower or 'probably' in pred_lower:
        return 'pred-p'
    elif pred_lower.startswith('b') or 'benign' in pred_lower:
        return 'pred-b'
    elif pred_lower.startswith('n') or 'neutral' in pred_lower:
        return 'pred-n'
    return ''


class VariantReportGenerator:
    """Generate HTML reports from Excel variant data"""

    def __init__(self, excel_path: str, output_dir: str = None):
        self.excel_path = Path(excel_path)
        self.output_dir = Path(output_dir) if output_dir else self.excel_path.parent / 'html_reports'
        self.sample_name = self.excel_path.stem
        self.variants = []
        self.headers = []

    def load_data(self):
        """Load data from Excel file"""
        wb = openpyxl.load_workbook(self.excel_path, read_only=True)
        ws = wb.worksheets[0]  # First worksheet (.filter)

        # Check if first row contains column numbers (integers) instead of headers
        # Some files have an extra row with column numbers before the header row
        first_row = [cell.value for cell in ws[1]]
        header_row = 1
        data_start_row = 2

        # If first cell is an integer (column number), headers are in row 2
        if isinstance(first_row[0], int):
            header_row = 2
            data_start_row = 3
            self.headers = [''] + [cell.value for cell in ws[2]]  # 1-indexed to match column numbers
        else:
            self.headers = [''] + first_row  # 1-indexed to match column numbers

        # Get variants
        for row in ws.iter_rows(min_row=data_start_row):
            if row[0].value is not None:
                variant = [''] + [cell.value for cell in row]  # 1-indexed
                self.variants.append(variant)

        wb.close()
        print(f"Loaded {len(self.variants)} variants from {self.excel_path.name}")

    def generate_all(self):
        """Generate single HTML file with dashboard and embedded variant pages"""
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Generate single embedded HTML file
        self.generate_embedded_report()

        print(f"Generated HTML report: {self.output_dir / f'{self.sample_name}.html'}")

    def generate_embedded_report(self):
        """Generate single HTML file with dashboard and embedded variant detail pages"""
        # Collect stats
        genes = set()
        tiers = {'tier-i': 0, 'tier-ii': 0, 'tier-iii': 0, 'tier-iv': 0, 'unknown': 0}

        for variant in self.variants:
            genes.add(variant[6] if len(variant) > 6 else 'Unknown')
            _, tier_class = parse_cancervar(variant[18] if len(variant) > 18 else '')
            tiers[tier_class] = tiers.get(tier_class, 0) + 1

        # Additional CSS for embedded view switching
        embedded_css = """
/* View switching */
.view-section {
    display: none;
}

.view-section.active {
    display: block;
}

/* Variant detail header with back button */
.variant-detail-header {
    display: flex;
    align-items: center;
    gap: 15px;
    margin-bottom: 20px;
}

.back-btn {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    padding: 10px 18px;
    background: white;
    border: 1px solid var(--border-color);
    border-radius: 6px;
    color: var(--primary-color);
    text-decoration: none;
    font-size: 0.9rem;
    font-weight: 500;
    cursor: pointer;
    transition: all 0.2s;
}

.back-btn:hover {
    background: var(--secondary-color);
    color: white;
    border-color: var(--secondary-color);
}

.back-btn svg {
    width: 16px;
    height: 16px;
}

/* View Details link styling */
.view-details-link {
    color: var(--secondary-color);
    text-decoration: none;
    cursor: pointer;
    font-weight: 500;
}

.view-details-link:hover {
    text-decoration: underline;
}

/* Navigation between variants */
.variant-nav {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-top: 20px;
    padding-top: 20px;
    border-top: 1px solid var(--border-color);
}

.nav-btn {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    padding: 10px 18px;
    background: var(--light-bg);
    border: 1px solid var(--border-color);
    border-radius: 6px;
    color: var(--primary-color);
    text-decoration: none;
    font-size: 0.9rem;
    font-weight: 500;
    cursor: pointer;
    transition: all 0.2s;
}

.nav-btn:hover {
    background: var(--secondary-color);
    color: white;
    border-color: var(--secondary-color);
}

.nav-btn.disabled {
    opacity: 0.4;
    cursor: not-allowed;
    pointer-events: none;
}

.nav-btn svg {
    width: 16px;
    height: 16px;
}

.variant-counter {
    font-size: 0.9rem;
    color: var(--text-muted);
}
"""

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Variant Report - {escape(self.sample_name)}</title>
    <style>
{get_css_styles()}
{embedded_css}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Variant Analysis Report</h1>
            <div class="subtitle">Sample: {escape(self.sample_name)}</div>
        </div>

        <!-- Dashboard View -->
        <div id="dashboard-view" class="view-section active">
            <div class="dashboard-stats">
                <div class="stat-card clickable" data-action="clear-filters" title="Click to clear all filters">
                    <div class="stat-value" id="visibleCount">{len(self.variants)}</div>
                    <div class="stat-label">Total Variants</div>
                </div>
                <div class="stat-card clickable" data-action="focus-gene" title="Click to focus on gene filter">
                    <div class="stat-value" style="color: var(--secondary-color);">{len(genes)}</div>
                    <div class="stat-label">Genes Affected</div>
                </div>
                <div class="stat-card clickable" data-tier="tier-i" title="Click to filter Tier I variants">
                    <div class="stat-value" style="color: var(--tier1-color);">{tiers.get('tier-i', 0)}</div>
                    <div class="stat-label">Tier I (Strong)</div>
                </div>
                <div class="stat-card clickable" data-tier="tier-ii" title="Click to filter Tier II variants">
                    <div class="stat-value" style="color: var(--tier2-color);">{tiers.get('tier-ii', 0)}</div>
                    <div class="stat-label">Tier II (Potential)</div>
                </div>
            </div>

            <div class="filter-bar">
                <div class="filter-group">
                    <label class="filter-label">Search</label>
                    <input type="text" id="searchInput" class="filter-input" placeholder="Search variants...">
                </div>
                <div class="filter-group">
                    <label class="filter-label">Gene</label>
                    <select id="geneFilter" class="filter-select">
                        <option value="">All Genes</option>
                        {''.join(f'<option value="{escape(str(g))}">{escape(str(g))}</option>' for g in sorted(genes))}
                    </select>
                </div>
                <div class="filter-group">
                    <label class="filter-label">Classification</label>
                    <select id="tierFilter" class="filter-select">
                        <option value="">All Classifications</option>
                        <option value="tier-i">Tier I - Strong</option>
                        <option value="tier-ii">Tier II - Potential</option>
                        <option value="tier-iii">Tier III - Unknown</option>
                        <option value="tier-iv">Tier IV - Benign</option>
                    </select>
                </div>
            </div>

            <table class="variant-table">
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Gene</th>
                        <th>Variant</th>
                        <th>HGVSc</th>
                        <th>HGVSp</th>
                        <th>VAF</th>
                        <th>Classification</th>
                        <th>Action</th>
                    </tr>
                </thead>
                <tbody>
"""

        for idx, variant in enumerate(self.variants):
            gene = escape(str(variant[6])) if len(variant) > 6 and variant[6] else '-'
            chr_pos = f"{variant[1]}:{variant[2]}" if len(variant) > 2 else '-'
            ref_alt = f"{variant[3]}>{variant[4]}" if len(variant) > 4 else '-'
            hgvsc = escape(str(variant[9])) if len(variant) > 9 and variant[9] else '-'
            hgvsp = escape(str(variant[10])) if len(variant) > 10 and variant[10] else '-'
            vaf = f"{float(variant[14])*100:.1f}%" if len(variant) > 14 and variant[14] else '-'

            tier_text, tier_class = parse_cancervar(variant[18] if len(variant) > 18 else '')

            html += f"""
                    <tr class="variant-row" data-gene="{gene}" data-tier="{tier_class}">
                        <td>{idx + 1}</td>
                        <td class="gene">{gene}</td>
                        <td style="font-family: monospace;">{escape(chr_pos)} {escape(ref_alt)}</td>
                        <td style="font-family: monospace;">{hgvsc}</td>
                        <td style="font-family: monospace;">{hgvsp}</td>
                        <td>{vaf}</td>
                        <td><span class="classification-badge {tier_class}">{escape(tier_text)}</span></td>
                        <td><a class="view-details-link" onclick="showVariant({idx + 1})">View Details</a></td>
                    </tr>
"""

        html += """
                </tbody>
            </table>
        </div>

"""

        # Generate embedded variant detail sections
        for idx, variant in enumerate(self.variants):
            html += self._generate_embedded_variant_section(idx, variant, len(self.variants))

        # Close container and add JavaScript
        html += """
    </div>
    <script>
""" + get_javascript() + """

// View switching functions
function showVariant(idx) {
    // Hide dashboard
    document.getElementById('dashboard-view').classList.remove('active');

    // Hide all variant views
    document.querySelectorAll('.variant-view').forEach(v => v.classList.remove('active'));

    // Show selected variant
    document.getElementById('variant-view-' + idx).classList.add('active');

    // Scroll to top
    window.scrollTo(0, 0);
}

function showDashboard() {
    // Hide all variant views
    document.querySelectorAll('.variant-view').forEach(v => v.classList.remove('active'));

    // Show dashboard
    document.getElementById('dashboard-view').classList.add('active');

    // Scroll to top
    window.scrollTo(0, 0);
}

function navigateVariant(currentIdx, direction) {
    const totalVariants = document.querySelectorAll('.variant-view').length;
    let newIdx = currentIdx + direction;

    if (newIdx >= 1 && newIdx <= totalVariants) {
        showVariant(newIdx);
    }
}
    </script>
</body>
</html>
"""

        with open(self.output_dir / f'{self.sample_name}.html', 'w', encoding='utf-8') as f:
            f.write(html)

    def _generate_embedded_variant_section(self, idx: int, variant: List, total_variants: int) -> str:
        """Generate embedded variant detail section for single-file report"""
        gene = str(variant[6]) if len(variant) > 6 and variant[6] else 'Unknown'
        hgvsc = str(variant[9]) if len(variant) > 9 and variant[9] else ''
        hgvsp = str(variant[10]) if len(variant) > 10 and variant[10] else ''
        tier_text, tier_class = parse_cancervar(variant[18] if len(variant) > 18 else '')

        # Build variant notation
        variant_notation = f"{hgvsc}"
        if hgvsp:
            variant_notation += f" ({hgvsp})"

        # Navigation button states
        prev_disabled = 'disabled' if idx == 0 else ''
        next_disabled = 'disabled' if idx == len(self.variants) - 1 else ''

        html = f"""
        <!-- Variant {idx + 1} Detail View -->
        <div id="variant-view-{idx + 1}" class="view-section variant-view">
            <div class="variant-detail-header">
                <button class="back-btn" onclick="showDashboard()">
                    <svg fill="none" stroke="currentColor" viewBox="0 0 24 24">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M10 19l-7-7m0 0l7-7m-7 7h18"/>
                    </svg>
                    Back to Dashboard
                </button>
            </div>

            <div class="gene-hero">
                <div>
                    <h1 class="gene-name">{escape(gene)}</h1>
                    <div class="variant-notation">{escape(variant_notation)}</div>
                </div>
                <div>
                    <span class="classification-badge {tier_class}">{escape(tier_text)}</span>
                </div>
            </div>
"""

        # Add all 6 panels
        html += self._generate_basic_info_panel(variant)
        html += self._generate_igv_screenshot_panel(variant)
        html += self._generate_sample_comparison_panel(variant)
        html += self._generate_additional_info_panel(variant)
        html += self._generate_population_panel(variant)
        html += self._generate_computational_panel(variant)

        # Add navigation between variants
        html += f"""
            <div class="variant-nav">
                <button class="nav-btn {prev_disabled}" onclick="navigateVariant({idx + 1}, -1)">
                    <svg fill="none" stroke="currentColor" viewBox="0 0 24 24">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M15 19l-7-7 7-7"/>
                    </svg>
                    Previous
                </button>
                <span class="variant-counter">Variant {idx + 1} of {total_variants}</span>
                <button class="nav-btn {next_disabled}" onclick="navigateVariant({idx + 1}, 1)">
                    Next
                    <svg fill="none" stroke="currentColor" viewBox="0 0 24 24">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 5l7 7-7 7"/>
                    </svg>
                </button>
            </div>
        </div>
"""
        return html

    def generate_dashboard(self):
        """Generate the dashboard landing page"""
        # Collect stats
        genes = set()
        tiers = {'tier-i': 0, 'tier-ii': 0, 'tier-iii': 0, 'tier-iv': 0, 'unknown': 0}

        for variant in self.variants:
            genes.add(variant[6] if len(variant) > 6 else 'Unknown')
            _, tier_class = parse_cancervar(variant[18] if len(variant) > 18 else '')
            tiers[tier_class] = tiers.get(tier_class, 0) + 1

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Variant Report - {escape(self.sample_name)}</title>
    <style>{get_css_styles()}</style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Variant Analysis Report</h1>
            <div class="subtitle">Sample: {escape(self.sample_name)}</div>
        </div>

        <div class="dashboard-stats">
            <div class="stat-card clickable" data-action="clear-filters" title="Click to clear all filters">
                <div class="stat-value" id="visibleCount">{len(self.variants)}</div>
                <div class="stat-label">Total Variants</div>
            </div>
            <div class="stat-card clickable" data-action="focus-gene" title="Click to focus on gene filter">
                <div class="stat-value" style="color: var(--secondary-color);">{len(genes)}</div>
                <div class="stat-label">Genes Affected</div>
            </div>
            <div class="stat-card clickable" data-tier="tier-i" title="Click to filter Tier I variants">
                <div class="stat-value" style="color: var(--tier1-color);">{tiers.get('tier-i', 0)}</div>
                <div class="stat-label">Tier I (Strong)</div>
            </div>
            <div class="stat-card clickable" data-tier="tier-ii" title="Click to filter Tier II variants">
                <div class="stat-value" style="color: var(--tier2-color);">{tiers.get('tier-ii', 0)}</div>
                <div class="stat-label">Tier II (Potential)</div>
            </div>
        </div>

        <div class="filter-bar">
            <div class="filter-group">
                <label class="filter-label">Search</label>
                <input type="text" id="searchInput" class="filter-input" placeholder="Search variants...">
            </div>
            <div class="filter-group">
                <label class="filter-label">Gene</label>
                <select id="geneFilter" class="filter-select">
                    <option value="">All Genes</option>
                    {''.join(f'<option value="{escape(str(g))}">{escape(str(g))}</option>' for g in sorted(genes))}
                </select>
            </div>
            <div class="filter-group">
                <label class="filter-label">Classification</label>
                <select id="tierFilter" class="filter-select">
                    <option value="">All Classifications</option>
                    <option value="tier-i">Tier I - Strong</option>
                    <option value="tier-ii">Tier II - Potential</option>
                    <option value="tier-iii">Tier III - Unknown</option>
                    <option value="tier-iv">Tier IV - Benign</option>
                </select>
            </div>
        </div>

        <table class="variant-table">
            <thead>
                <tr>
                    <th>#</th>
                    <th>Gene</th>
                    <th>Variant</th>
                    <th>HGVSc</th>
                    <th>HGVSp</th>
                    <th>VAF</th>
                    <th>Classification</th>
                    <th>Action</th>
                </tr>
            </thead>
            <tbody>
"""

        for idx, variant in enumerate(self.variants):
            gene = escape(str(variant[6])) if len(variant) > 6 and variant[6] else '-'
            chr_pos = f"{variant[1]}:{variant[2]}" if len(variant) > 2 else '-'
            ref_alt = f"{variant[3]}>{variant[4]}" if len(variant) > 4 else '-'
            hgvsc = escape(str(variant[9])) if len(variant) > 9 and variant[9] else '-'
            hgvsp = escape(str(variant[10])) if len(variant) > 10 and variant[10] else '-'
            vaf = f"{float(variant[14])*100:.1f}%" if len(variant) > 14 and variant[14] else '-'

            tier_text, tier_class = parse_cancervar(variant[18] if len(variant) > 18 else '')

            html += f"""
                <tr class="variant-row" data-gene="{gene}" data-tier="{tier_class}">
                    <td>{idx + 1}</td>
                    <td class="gene">{gene}</td>
                    <td style="font-family: monospace;">{escape(chr_pos)} {escape(ref_alt)}</td>
                    <td style="font-family: monospace;">{hgvsc}</td>
                    <td style="font-family: monospace;">{hgvsp}</td>
                    <td>{vaf}</td>
                    <td><span class="classification-badge {tier_class}">{escape(tier_text)}</span></td>
                    <td><a href="variant_{idx + 1}.html">View Details</a></td>
                </tr>
"""

        html += """
            </tbody>
        </table>
    </div>
    <script>
""" + get_javascript() + """
    </script>
</body>
</html>
"""

        with open(self.output_dir / f'{self.sample_name}.html', 'w', encoding='utf-8') as f:
            f.write(html)

    def generate_variant_page(self, idx: int, variant: List):
        """Generate individual variant page"""
        gene = str(variant[6]) if len(variant) > 6 and variant[6] else 'Unknown'
        hgvsc = str(variant[9]) if len(variant) > 9 and variant[9] else ''
        hgvsp = str(variant[10]) if len(variant) > 10 and variant[10] else ''
        tier_text, tier_class = parse_cancervar(variant[18] if len(variant) > 18 else '')

        # Build variant notation
        variant_notation = f"{hgvsc}"
        if hgvsp:
            variant_notation += f" ({hgvsp})"

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{escape(gene)} - {escape(hgvsc)} | Variant Report</title>
    <style>{get_css_styles()}</style>
</head>
<body>
    <div class="container">
        <div class="breadcrumb">
            <a href="{self.sample_name}.html">Dashboard</a> &rarr; Variant {idx + 1}: {escape(gene)}
        </div>

        <div class="gene-hero">
            <div>
                <h1 class="gene-name">{escape(gene)}</h1>
                <div class="variant-notation">{escape(variant_notation)}</div>
            </div>
            <div>
                <span class="classification-badge {tier_class}">{escape(tier_text)}</span>
            </div>
        </div>
"""

        # Panel 1: Basic Variant Information (cols 1-18)
        html += self._generate_basic_info_panel(variant)

        # Panel 2: IGV Screenshot
        html += self._generate_igv_screenshot_panel(variant)

        # Panel 3: Sample Comparison (cols 19-22)
        html += self._generate_sample_comparison_panel(variant)

        # Panel 4: Additional Variant Information (cols 23-31)
        html += self._generate_additional_info_panel(variant)

        # Panel 5: Population Databases (cols 32-89)
        html += self._generate_population_panel(variant)

        # Panel 6: Computational Scores (cols 90-154)
        html += self._generate_computational_panel(variant)

        html += """
    </div>
    <script>
""" + get_javascript() + """
    </script>
</body>
</html>
"""

        with open(self.output_dir / f'variant_{idx + 1}.html', 'w', encoding='utf-8') as f:
            f.write(html)

    def generate_single_page(self, output_path: str = None):
        """Generate a single HTML page containing all variants"""
        if output_path is None:
            output_path = self.output_dir / f'{self.sample_name}.html'
        else:
            output_path = Path(output_path)

        # Ensure parent directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Collect stats
        genes = set()
        tiers = {'tier-i': 0, 'tier-ii': 0, 'tier-iii': 0, 'tier-iv': 0, 'unknown': 0}

        for variant in self.variants:
            genes.add(variant[6] if len(variant) > 6 else 'Unknown')
            _, tier_class = parse_cancervar(variant[18] if len(variant) > 18 else '')
            tiers[tier_class] = tiers.get(tier_class, 0) + 1

        # Additional CSS for single-page layout
        single_page_css = """
/* Single Page Variant Cards */
.variant-card {
    background: white;
    border-radius: 12px;
    margin-bottom: 25px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.1);
    overflow: hidden;
    border: 1px solid var(--border-color);
}

.variant-card-header {
    background: linear-gradient(135deg, var(--primary-color) 0%, #34495e 100%);
    color: white;
    padding: 18px 25px;
    cursor: pointer;
    display: flex;
    justify-content: space-between;
    align-items: center;
    transition: background 0.2s;
}

.variant-card-header:hover {
    background: linear-gradient(135deg, #34495e 0%, #2c3e50 100%);
}

.variant-card-header .gene-name {
    font-size: 1.8rem;
    font-weight: 700;
    color: white;
    margin: 0;
}

.variant-card-header .variant-info {
    font-size: 0.95rem;
    opacity: 0.9;
    font-family: 'Monaco', 'Menlo', monospace;
    margin-top: 4px;
}

.variant-card-header .header-right {
    display: flex;
    align-items: center;
    gap: 15px;
}

.variant-card-header .card-toggle {
    font-size: 1.5rem;
    transition: transform 0.3s;
}

.variant-card.collapsed .card-toggle {
    transform: rotate(-90deg);
}

.variant-card.collapsed .variant-card-content {
    display: none;
}

.variant-card-content {
    padding: 20px;
    background: var(--light-bg);
}

/* Quick Nav */
.quick-nav {
    background: white;
    border-radius: 10px;
    padding: 15px 20px;
    margin-bottom: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}

.quick-nav-title {
    font-size: 0.8rem;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 10px;
}

.quick-nav-links {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
}

.quick-nav-link {
    display: inline-block;
    padding: 6px 14px;
    background: var(--light-bg);
    border-radius: 20px;
    color: var(--primary-color);
    text-decoration: none;
    font-size: 0.85rem;
    font-weight: 500;
    transition: all 0.2s;
    border: 1px solid var(--border-color);
}

.quick-nav-link:hover {
    background: var(--secondary-color);
    color: white;
    border-color: var(--secondary-color);
}

.quick-nav-link .tier-dot {
    display: inline-block;
    width: 8px;
    height: 8px;
    border-radius: 50%;
    margin-right: 6px;
}

.quick-nav-link .tier-dot.tier-i { background: var(--tier1-color); }
.quick-nav-link .tier-dot.tier-ii { background: var(--tier2-color); }
.quick-nav-link .tier-dot.tier-iii { background: var(--tier3-color); }
.quick-nav-link .tier-dot.tier-iv { background: var(--tier4-color); }
.quick-nav-link .tier-dot.unknown { background: var(--text-muted); }

/* Expand/Collapse All */
.controls-bar {
    display: flex;
    gap: 10px;
    margin-bottom: 15px;
}

.control-btn {
    padding: 8px 16px;
    background: var(--secondary-color);
    color: white;
    border: none;
    border-radius: 6px;
    cursor: pointer;
    font-size: 0.85rem;
    font-weight: 500;
    transition: background 0.2s;
}

.control-btn:hover {
    background: #2980b9;
}
"""

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Variant Report - {escape(self.sample_name)}</title>
    <style>
{get_css_styles()}
{single_page_css}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Variant Analysis Report</h1>
            <div class="subtitle">Sample: {escape(self.sample_name)}</div>
        </div>

        <div class="dashboard-stats">
            <div class="stat-card">
                <div class="stat-value">{len(self.variants)}</div>
                <div class="stat-label">Total Variants</div>
            </div>
            <div class="stat-card">
                <div class="stat-value" style="color: var(--secondary-color);">{len(genes)}</div>
                <div class="stat-label">Genes Affected</div>
            </div>
            <div class="stat-card">
                <div class="stat-value" style="color: var(--tier1-color);">{tiers.get('tier-i', 0)}</div>
                <div class="stat-label">Tier I (Strong)</div>
            </div>
            <div class="stat-card">
                <div class="stat-value" style="color: var(--tier2-color);">{tiers.get('tier-ii', 0)}</div>
                <div class="stat-label">Tier II (Potential)</div>
            </div>
        </div>

        <div class="quick-nav">
            <div class="quick-nav-title">Quick Navigation</div>
            <div class="quick-nav-links">
"""

        # Add quick nav links for each variant
        for idx, variant in enumerate(self.variants):
            gene = str(variant[6]) if len(variant) > 6 and variant[6] else 'Unknown'
            hgvsp = str(variant[10]) if len(variant) > 10 and variant[10] else ''
            _, tier_class = parse_cancervar(variant[18] if len(variant) > 18 else '')
            label = f"{gene}"
            if hgvsp:
                label += f" {hgvsp}"

            html += f"""
                <a href="#variant-{idx + 1}" class="quick-nav-link">
                    <span class="tier-dot {tier_class}"></span>{escape(label)}
                </a>
"""

        html += """
            </div>
        </div>

        <div class="controls-bar">
            <button class="control-btn" onclick="expandAll()">Expand All</button>
            <button class="control-btn" onclick="collapseAll()">Collapse All</button>
        </div>
"""

        # Generate each variant card
        for idx, variant in enumerate(self.variants):
            gene = str(variant[6]) if len(variant) > 6 and variant[6] else 'Unknown'
            hgvsc = str(variant[9]) if len(variant) > 9 and variant[9] else ''
            hgvsp = str(variant[10]) if len(variant) > 10 and variant[10] else ''
            tier_text, tier_class = parse_cancervar(variant[18] if len(variant) > 18 else '')

            variant_notation = f"{hgvsc}"
            if hgvsp:
                variant_notation += f" ({hgvsp})"

            html += f"""
        <div class="variant-card" id="variant-{idx + 1}">
            <div class="variant-card-header" onclick="toggleVariantCard(this)">
                <div>
                    <h2 class="gene-name">{escape(gene)}</h2>
                    <div class="variant-info">{escape(variant_notation)}</div>
                </div>
                <div class="header-right">
                    <span class="classification-badge {tier_class}">{escape(tier_text)}</span>
                    <span class="card-toggle">&#9662;</span>
                </div>
            </div>
            <div class="variant-card-content">
"""

            # Add all 6 panels
            html += self._generate_basic_info_panel(variant)
            html += self._generate_igv_screenshot_panel(variant)
            html += self._generate_sample_comparison_panel(variant)
            html += self._generate_additional_info_panel(variant)
            html += self._generate_population_panel(variant)
            html += self._generate_computational_panel(variant)

            html += """
            </div>
        </div>
"""

        # Close container and add JavaScript
        html += """
    </div>
    <script>
""" + get_javascript() + """

// Toggle individual variant card
function toggleVariantCard(header) {
    const card = header.parentElement;
    card.classList.toggle('collapsed');
}

// Expand all variant cards
function expandAll() {
    document.querySelectorAll('.variant-card').forEach(card => {
        card.classList.remove('collapsed');
    });
}

// Collapse all variant cards
function collapseAll() {
    document.querySelectorAll('.variant-card').forEach(card => {
        card.classList.add('collapsed');
    });
}

// Smooth scroll to variant when clicking quick nav
document.querySelectorAll('.quick-nav-link').forEach(link => {
    link.addEventListener('click', (e) => {
        e.preventDefault();
        const targetId = link.getAttribute('href').substring(1);
        const target = document.getElementById(targetId);
        if (target) {
            // Expand the card if collapsed
            target.classList.remove('collapsed');
            // Smooth scroll
            target.scrollIntoView({ behavior: 'smooth', block: 'start' });
        }
    });
});
    </script>
</body>
</html>
"""

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)

        print(f"Generated single-page report: {output_path}")

    def _generate_basic_info_panel(self, variant: List) -> str:
        """Generate Basic Variant Information panel"""
        html = """
        <div class="panel panel-basic">
            <div class="panel-header">
                <h3 class="panel-title">Basic Variant Information</h3>
                <span class="panel-toggle">&#9662;</span>
            </div>
            <div class="panel-content">
                <div class="data-grid">
"""

        # Key fields with larger display
        key_fields = [
            (1, 'Chromosome', False),
            (2, 'Position', False),
            (3, 'Reference', True),
            (4, 'Alternate', True),
            (5, 'Genotype', False),
            (6, 'Gene', False),
            (7, 'Transcript', True),
            (8, 'HGVSg', True),
            (9, 'HGVSc', True),
            (10, 'HGVSp', True),
            (11, 'Allele Depth', False),
            (12, 'Read Depth', False),
            (13, 'Quality', False),
            (14, 'VAF', False),
            (15, 'dbSNP', True),
            (16, 'Full Notation', True),
            (17, 'COSMIC', True),
            (18, 'CancerVar', False),
        ]

        for col, label, is_mono in key_fields:
            val = variant[col] if len(variant) > col else None
            formatted = format_value(val)

            # Special formatting for VAF
            if col == 14 and val and val != '.':
                try:
                    formatted = f"{float(val)*100:.2f}%"
                except:
                    pass

            mono_class = ' monospace' if is_mono else ''
            large_class = ' large' if col == 6 else ''

            html += f"""
                    <div class="data-item">
                        <div class="data-label">{escape(label)}</div>
                        <div class="data-value{mono_class}{large_class}">{escape(formatted)}</div>
                    </div>
"""

        html += """
                </div>
            </div>
        </div>
"""
        return html

    def _generate_igv_screenshot_panel(self, variant: List, snapshot_dir: str = "../SnapShots") -> str:
        """Generate IGV Screenshot panel"""
        gene = str(variant[6]) if len(variant) > 6 and variant[6] else 'Unknown'
        position = str(variant[2]) if len(variant) > 2 and variant[2] else ''

        # Extract short sample ID (e.g., AML-452 from AML-452-KHK-TMSP_S1)
        parts = self.sample_name.split('-')
        short_sample = '-'.join(parts[:2]) if len(parts) >= 2 else self.sample_name

        # Build screenshot filename (matches processVCF.sh naming convention: short_sample-gene-position.png)
        if gene and position:
            screenshot_filename = f"{short_sample}-{gene}-{position}.png"
            screenshot_path = f"{snapshot_dir}/{screenshot_filename}"
        else:
            screenshot_filename = None
            screenshot_path = None

        html = """
        <div class="panel panel-igv collapsed">
            <div class="panel-header">
                <h3 class="panel-title">IGV Screenshot</h3>
                <span class="panel-toggle">&#9662;</span>
            </div>
            <div class="panel-content">
"""

        if screenshot_path:
            html += f"""
                <img src="{escape(screenshot_path)}" alt="IGV Screenshot for {escape(gene)} at position {escape(position)}"
                     class="igv-image" onerror="this.style.display='none'; this.nextElementSibling.style.display='block';">
                <div class="igv-no-image" style="display: none;">
                    Screenshot not available: {escape(screenshot_filename)}
                </div>
"""
        else:
            html += """
                <div class="igv-no-image">
                    No IGV screenshot available for this variant.
                </div>
"""

        html += """
            </div>
        </div>
"""
        return html

    def _generate_sample_comparison_panel(self, variant: List) -> str:
        """Generate Sample Comparison panel"""
        html = """
        <div class="panel panel-comparison">
            <div class="panel-header">
                <h3 class="panel-title">Sample Comparison Data</h3>
                <span class="panel-toggle">&#9662;</span>
            </div>
            <div class="panel-content">
                <div class="data-grid">
"""

        for col in ColumnGroups.SAMPLE_COMPARISON:
            if col < len(self.headers) and col < len(variant):
                label = self.headers[col] if self.headers[col] else f'Column {col}'
                val = format_value(variant[col])
                html += f"""
                    <div class="data-item">
                        <div class="data-label">{escape(str(label))}</div>
                        <div class="data-value">{escape(val)}</div>
                    </div>
"""

        html += """
                </div>
            </div>
        </div>
"""
        return html

    def _generate_additional_info_panel(self, variant: List) -> str:
        """Generate Additional Variant Information panel"""
        html = """
        <div class="panel panel-additional">
            <div class="panel-header">
                <h3 class="panel-title">Additional Variant Information</h3>
                <span class="panel-toggle">&#9662;</span>
            </div>
            <div class="panel-content">
                <div class="data-grid">
"""

        for col in ColumnGroups.ADDITIONAL_INFO:
            if col < len(self.headers) and col < len(variant):
                label = self.headers[col] if self.headers[col] else f'Column {col}'
                val = format_value(variant[col])
                html += f"""
                    <div class="data-item">
                        <div class="data-label">{escape(str(label))}</div>
                        <div class="data-value monospace">{escape(val)}</div>
                    </div>
"""

        html += """
                </div>
            </div>
        </div>
"""
        return html

    def _generate_population_panel(self, variant: List) -> str:
        """Generate Population Databases panel grouped by database"""
        html = """
        <div class="panel panel-population">
            <div class="panel-header">
                <h3 class="panel-title">Population Frequency Databases</h3>
                <span class="panel-toggle">&#9662;</span>
            </div>
            <div class="panel-content">
"""

        for db_name, cols in POP_DB_GROUPS.items():
            # Check if any values in this group are non-empty
            has_data = False
            for col in cols:
                if col < len(variant) and not is_empty_value(variant[col]):
                    has_data = True
                    break

            if not has_data:
                continue

            html += f"""
                <div class="pop-db-section">
                    <div class="pop-db-title">{escape(db_name)}</div>
                    <div class="data-grid">
"""

            for col in cols:
                if col < len(self.headers) and col < len(variant):
                    if is_empty_value(variant[col]):
                        continue

                    label = self.headers[col] if self.headers[col] else f'Column {col}'
                    val = variant[col]
                    formatted = format_value(val)

                    # Add frequency bar for numeric frequencies
                    freq_bar = ''
                    try:
                        freq = float(val)
                        if 0 <= freq <= 1:
                            bar_width = min(freq * 100 * 10, 100)  # Scale up small frequencies
                            bar_class = get_freq_bar_class(freq)
                            freq_bar = f'''
                        <div class="freq-bar-container">
                            <div class="freq-bar {bar_class}" style="width: {bar_width}%"></div>
                        </div>
'''
                    except:
                        pass

                    html += f"""
                        <div class="data-item">
                            <div class="data-label">{escape(str(label))}</div>
                            <div class="data-value monospace">{escape(formatted)}</div>
                            {freq_bar}
                        </div>
"""

            html += """
                    </div>
                </div>
"""

        html += """
            </div>
        </div>
"""
        return html

    def _generate_computational_panel(self, variant: List) -> str:
        """Generate Computational Scores panel with ACMG chips and score table"""
        html = """
        <div class="panel panel-computational">
            <div class="panel-header">
                <h3 class="panel-title">Computational Predictions & ACMG Criteria</h3>
                <span class="panel-toggle">&#9662;</span>
            </div>
            <div class="panel-content">
"""

        # InterVar classification
        intervar = format_value(variant[90]) if len(variant) > 90 else '-'
        html += f"""
                <div style="margin-bottom: 20px;">
                    <div class="data-label">InterVar Classification</div>
                    <div class="data-value large">{escape(intervar)}</div>
                </div>
"""

        # ACMG Criteria chips
        html += """
                <div style="margin-bottom: 25px;">
                    <div class="data-label" style="margin-bottom: 10px;">ACMG Criteria</div>
"""

        for category, criteria in ACMG_CRITERIA.items():
            html += f"""
                    <div style="margin-bottom: 8px;">
                        <small style="color: var(--text-muted);">{escape(category)}</small>
                        <div class="acmg-grid">
"""

            for col, name in criteria:
                val = variant[col] if len(variant) > col else None
                is_active = val and val != '.' and val != '0' and str(val).strip() != ''

                # Determine chip class
                chip_class = 'inactive'
                if is_active:
                    if name.startswith('PVS'):
                        chip_class = 'pvs'
                    elif name.startswith('PS'):
                        chip_class = 'ps'
                    elif name.startswith('PM'):
                        chip_class = 'pm'
                    elif name.startswith('PP'):
                        chip_class = 'pp'
                    elif name.startswith('BA'):
                        chip_class = 'ba'
                    elif name.startswith('BS'):
                        chip_class = 'bs'
                    elif name.startswith('BP'):
                        chip_class = 'bp'

                html += f"""
                            <span class="acmg-chip {chip_class}">{escape(name)}</span>
"""

            html += """
                        </div>
                    </div>
"""

        html += """
                </div>
"""

        # ClinVar Information
        html += """
                <div style="margin-bottom: 25px;">
                    <div class="data-label" style="margin-bottom: 10px;">ClinVar Information</div>
                    <div class="data-grid">
"""

        clinvar_labels = ['Allele ID', 'Disease Name', 'Disease Database', 'Review Status', 'Clinical Significance']
        for i, col in enumerate(CLINVAR_COLS):
            if col < len(variant):
                label = clinvar_labels[i] if i < len(clinvar_labels) else f'ClinVar {col}'
                val = format_value(variant[col])

                # Special handling for clinical significance
                badge_html = ''
                if col == 123 and variant[col] and variant[col] != '.':
                    sig_text, sig_class = parse_clinsig(variant[col])
                    badge_html = f'<span class="classification-badge {sig_class}" style="margin-top: 5px;">{escape(sig_text)}</span>'

                html += f"""
                        <div class="data-item">
                            <div class="data-label">{escape(label)}</div>
                            <div class="data-value">{escape(val)}</div>
                            {badge_html}
                        </div>
"""

        html += """
                    </div>
                </div>
"""

        # Prediction Scores Table
        html += """
                <div>
                    <div class="data-label" style="margin-bottom: 10px;">Prediction Scores</div>
                    <table class="score-table">
                        <thead>
                            <tr>
                                <th>Tool</th>
                                <th>Score</th>
                                <th>Prediction</th>
                            </tr>
                        </thead>
                        <tbody>
"""

        for tool_name, cols in PREDICTION_SCORES.items():
            if tool_name in ['InterVar', 'Conservation']:
                continue

            score_val = '-'
            pred_val = '-'
            pred_class = ''

            for col, col_name in cols:
                if col < len(variant):
                    val = variant[col]
                    if 'score' in col_name.lower() or col_name in ['MCAP', 'REVEL', 'CADD_raw', 'CADD_phred', 'GERP++_RS']:
                        score_val = format_value(val)
                    elif 'pred' in col_name.lower():
                        pred_val = format_value(val)
                        pred_class = get_prediction_class(val)

            # Skip if no data
            if score_val == '-' and pred_val == '-':
                continue

            pred_badge = f'<span class="pred-badge {pred_class}">{escape(pred_val)}</span>' if pred_class else escape(pred_val)

            html += f"""
                            <tr>
                                <td>{escape(tool_name)}</td>
                                <td>{escape(score_val)}</td>
                                <td>{pred_badge}</td>
                            </tr>
"""

        html += """
                        </tbody>
                    </table>
                </div>
"""

        # Conservation Scores
        html += """
                <div style="margin-top: 20px;">
                    <div class="data-label" style="margin-bottom: 10px;">Conservation Scores</div>
                    <div class="data-grid">
"""

        for col, col_name in PREDICTION_SCORES.get('Conservation', []) + PREDICTION_SCORES.get('phyloP', []) + PREDICTION_SCORES.get('SiPhy', []):
            if col < len(variant) and not is_empty_value(variant[col]):
                html += f"""
                        <div class="data-item">
                            <div class="data-label">{escape(col_name)}</div>
                            <div class="data-value monospace">{escape(format_value(variant[col]))}</div>
                        </div>
"""

        html += """
                    </div>
                </div>
            </div>
        </div>
"""
        return html


def generate_summary_html(html_dir: str, output_file: str = None):
    """Generate Summary.html dashboard linking to all sample reports"""
    html_path = Path(html_dir)

    if output_file is None:
        output_file = html_path / 'Summary.html'
    else:
        output_file = Path(output_file)

    # Find all sample HTML files (exclude Summary.html itself)
    sample_files = sorted([f for f in html_path.glob('*.html')
                          if f.name != 'Summary.html' and not f.name.startswith('variant_')])

    if not sample_files:
        print(f"No sample HTML files found in {html_dir}")
        return

    # Collect sample info
    samples = []
    for f in sample_files:
        sample_name = f.stem
        # Try to extract variant count from the file
        variant_count = 0
        try:
            with open(f, 'r', encoding='utf-8') as fp:
                content = fp.read()
                # Count variant-view divs
                variant_count = content.count('class="view-section variant-view"')
        except:
            pass

        samples.append({
            'name': sample_name,
            'file': f.name,
            'variants': variant_count
        })

    # Generate Summary HTML
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Variant Analysis Summary</title>
    <style>
:root {{
    --primary-color: #2c3e50;
    --secondary-color: #3498db;
    --accent-color: #e74c3c;
    --success-color: #27ae60;
    --light-bg: #f8f9fa;
    --border-color: #dee2e6;
    --text-muted: #6c757d;
}}

* {{
    box-sizing: border-box;
}}

body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
    line-height: 1.5;
    color: #212529;
    background-color: #f5f6fa;
    margin: 0;
    padding: 0;
}}

.container {{
    max-width: 1200px;
    margin: 0 auto;
    padding: 20px;
}}

.header {{
    background: linear-gradient(135deg, var(--primary-color) 0%, #34495e 100%);
    color: white;
    padding: 30px;
    margin-bottom: 30px;
    border-radius: 12px;
    box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    text-align: center;
}}

.header h1 {{
    margin: 0 0 10px 0;
    font-size: 2.2rem;
    font-weight: 700;
}}

.header .subtitle {{
    opacity: 0.9;
    font-size: 1rem;
}}

.stats-row {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 20px;
    margin-bottom: 30px;
}}

.stat-card {{
    background: white;
    border-radius: 12px;
    padding: 25px;
    text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}}

.stat-value {{
    font-size: 3rem;
    font-weight: 700;
    color: var(--primary-color);
}}

.stat-label {{
    font-size: 0.9rem;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-top: 5px;
}}

.section-title {{
    font-size: 1.3rem;
    font-weight: 600;
    color: var(--primary-color);
    margin-bottom: 20px;
    padding-bottom: 10px;
    border-bottom: 2px solid var(--border-color);
}}

.sample-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
    gap: 20px;
}}

.sample-card {{
    background: white;
    border-radius: 12px;
    padding: 25px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    transition: transform 0.2s, box-shadow 0.2s;
    text-decoration: none;
    color: inherit;
    display: block;
    border: 2px solid transparent;
}}

.sample-card:hover {{
    transform: translateY(-5px);
    box-shadow: 0 8px 25px rgba(0,0,0,0.15);
    border-color: var(--secondary-color);
}}

.sample-name {{
    font-size: 1.2rem;
    font-weight: 600;
    color: var(--primary-color);
    margin-bottom: 10px;
    word-break: break-word;
}}

.sample-info {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-top: 15px;
    padding-top: 15px;
    border-top: 1px solid var(--border-color);
}}

.variant-count {{
    font-size: 1.5rem;
    font-weight: 700;
    color: var(--secondary-color);
}}

.variant-label {{
    font-size: 0.8rem;
    color: var(--text-muted);
    text-transform: uppercase;
}}

.view-btn {{
    display: inline-block;
    padding: 8px 20px;
    background: var(--secondary-color);
    color: white;
    border-radius: 6px;
    font-size: 0.85rem;
    font-weight: 500;
    text-decoration: none;
}}

.sample-card:hover .view-btn {{
    background: #2980b9;
}}

/* QC Placeholder Section */
.qc-section {{
    background: white;
    border-radius: 12px;
    padding: 25px;
    margin-top: 30px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}}

.qc-placeholder {{
    text-align: center;
    padding: 40px;
    color: var(--text-muted);
    font-style: italic;
    background: var(--light-bg);
    border-radius: 8px;
    border: 2px dashed var(--border-color);
}}

.footer {{
    text-align: center;
    margin-top: 40px;
    padding: 20px;
    color: var(--text-muted);
    font-size: 0.85rem;
}}

@media (max-width: 768px) {{
    .container {{
        padding: 15px;
    }}

    .header {{
        padding: 20px;
    }}

    .header h1 {{
        font-size: 1.6rem;
    }}

    .sample-grid {{
        grid-template-columns: 1fr;
    }}
}}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Variant Analysis Summary</h1>
            <div class="subtitle">Generated: {Path(sample_files[0]).stat().st_mtime if sample_files else ''}</div>
        </div>

        <div class="stats-row">
            <div class="stat-card">
                <div class="stat-value">{len(samples)}</div>
                <div class="stat-label">Samples Analyzed</div>
            </div>
            <div class="stat-card">
                <div class="stat-value" style="color: var(--secondary-color);">{sum(s['variants'] for s in samples)}</div>
                <div class="stat-label">Total Variants</div>
            </div>
        </div>

        <div class="section-title">Sample Reports</div>
        <div class="sample-grid">
"""

    for sample in samples:
        html += f"""
            <a href="{escape(sample['file'])}" class="sample-card">
                <div class="sample-name">{escape(sample['name'])}</div>
                <div class="sample-info">
                    <div>
                        <div class="variant-count">{sample['variants']}</div>
                        <div class="variant-label">Variants</div>
                    </div>
                    <span class="view-btn">View Report</span>
                </div>
            </a>
"""

    html += """
        </div>

        <div class="qc-section">
            <div class="section-title">Quality Control</div>
            <div class="qc-placeholder">
                Sequencing QC and Coverage QC data will be displayed here in future updates.
            </div>
        </div>

        <div class="footer">
            Generated by processVCF Pipeline
        </div>
    </div>
</body>
</html>
"""

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"Generated Summary.html with {len(samples)} sample(s): {output_file}")


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='Convert Excel variant files to HTML reports',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate multi-page report (dashboard + individual variant pages)
  python3 excel_to_html_report.py sample.xlsx ./html_output

  # Generate single-page report (all variants on one page)
  python3 excel_to_html_report.py sample.xlsx --single-page

  # Generate single-page report to specific file
  python3 excel_to_html_report.py sample.xlsx --single-page -o report.html

  # Generate both formats
  python3 excel_to_html_report.py sample.xlsx ./html_output --single-page

  # Generate Summary.html landing page from existing HTML reports
  python3 excel_to_html_report.py --summary ./html_reports/
"""
    )

    parser.add_argument('excel_file', nargs='?', default=None,
                        help='Input Excel file (.xlsx)')
    parser.add_argument('output_dir', nargs='?', default=None,
                        help='Output directory for multi-page reports (default: ./html_reports)')
    parser.add_argument('--single-page', '-s', action='store_true',
                        help='Generate a single HTML page with all variants')
    parser.add_argument('--output', '-o', default=None,
                        help='Output file path for single-page report (default: <sample_name>.html)')
    parser.add_argument('--summary', metavar='HTML_DIR',
                        help='Generate Summary.html from existing HTML reports in the specified directory')

    args = parser.parse_args()

    # Handle --summary mode
    if args.summary:
        if not os.path.isdir(args.summary):
            print(f"Error: Directory not found: {args.summary}")
            sys.exit(1)
        generate_summary_html(args.summary)
        return

    # Normal mode requires excel_file
    if not args.excel_file:
        parser.print_help()
        sys.exit(1)

    if not os.path.exists(args.excel_file):
        print(f"Error: File not found: {args.excel_file}")
        sys.exit(1)

    generator = VariantReportGenerator(args.excel_file, args.output_dir)
    generator.load_data()

    if args.single_page:
        # Generate single-page report
        output_path = args.output
        if output_path is None:
            # Default to same directory as Excel file
            output_path = Path(args.excel_file).parent / f'{generator.sample_name}.html'
        generator.generate_single_page(output_path)

        # Also generate multi-page if output_dir was specified
        if args.output_dir:
            generator.generate_all()
            print(f"\nOpen {generator.output_dir / generator.sample_name}.html for multi-page view.")
    else:
        # Generate multi-page report only
        generator.generate_all()
        print(f"\nOpen {generator.output_dir / generator.sample_name}.html in a browser to view the report.")


if __name__ == '__main__':
    main()
